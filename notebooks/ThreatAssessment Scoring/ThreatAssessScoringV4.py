
def main():

    import sys
    import os
    import urllib3
    from configparser import ConfigParser

    # Add your local ThreatConnect SDK to path
    sys.path.append(r"Z:\HTOC\Data_Analytics\threatconnect")
    from ThreatConnect import ThreatConnect
    from RequestObject import RequestObject
    from Owners import Owners

    # Add your project repo to path
    project_root = r"C:\Users\jaskew\Documents\project_repository\scripts\Data Movement\ThrearConnect-api-pull"
    if project_root not in sys.path:
        sys.path.append(project_root)

    from utils.config_loader import load_config

    # Load API config
    config_path = os.path.join(project_root, "utils", "config.json")
    try:
        api_secret_key, api_access_id, api_base_url, api_default_org = load_config(config_path)
        display(f"Loaded config from: {config_path}")
        display(f"Base URL: {api_base_url}")
        display(f"Access ID: {api_access_id}")
        display(f"Default Org: {api_default_org}")
    except Exception as e:
        display(f"[ERROR] Failed to load configuration: {e}")
        sys.exit(1)

    # Disable SSL verification warnings (use cautiously)
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    verify_ssl = False

    # Initialize ThreatConnect session
    try:
        tc = ThreatConnect(api_access_id, api_secret_key, api_default_org, api_base_url)
        display("ThreatConnect initialized.")
    except Exception as e:
        display(f"[ERROR] Failed to initialize ThreatConnect: {e}")
        sys.exit(1)

    # Define the owner (organization scope)
    owner = 'HTOC Org'

    # Create a request object to fetch indicators (or other data)
    try:
        ro = RequestObject()
        ro.set_http_method('GET')
        ro.set_owner(owner)
        ro.set_owner_allowed(True)
        # ro.set_resource_pagination(True)  # Uncomment if needed
        display("RequestObject successfully created.")
    except Exception as e:
        display(f"[ERROR] Failed to initialize RequestObject: {e}")
        sys.exit(1)

    import pandas as pd
    from datetime import datetime, timedelta
    import pytz
    import urllib.parse

    # Configuration for ThreatConnect indicator query
    QUERY_LOOKBACK_DAYS = 360  # days of lastObserved activity to include
    INDICATOR_TYPE_NAMES = [
        "Address", "EmailAddress", "File", "Host", "URL", "ASN", "CIDR",
        "Email Subject", "Hashtag", "Mutex", "Registry Key", "User Agent",
    ]
    OWNER_NAMES = [
        'HTOC Org',
        'CISA Federal Feed',
        'CMS_CTI',
        'Crowdstrike Falcon Intelligence',
        'DHS CISCP',
        'Intel471',
        'Mandiant Advantage Threat Intelligence',
        'VA_TIP Data',
    ]
    RESULT_PAGE_SIZE = 500  # keep this smaller; same fields, just paged

    # Setup
    cutoff = pd.Timestamp.utcnow()
    start_date = (datetime.now(pytz.UTC) - timedelta(days=QUERY_LOOKBACK_DAYS)).date()
    start = f"{start_date}T00:00:00Z"

    type_names = INDICATOR_TYPE_NAMES
    type_name_condition = ", ".join([f'"{t}"' for t in type_names])

    list_of_owners = OWNER_NAMES

    # Build owner IN (...) clause
    owner_condition = ", ".join([f'"{o}"' for o in list_of_owners])

    tql_raw = (
        f'ownerName IN ({owner_condition}) AND '
        f'typeName IN ({type_name_condition}) AND '
        f'lastObserved >= "{start}"'
    )

    tql_encoded = urllib.parse.quote(tql_raw)

    final_results = []

    # Query indicators (paginate so you don't 502 with heavy fields)
    # Create a NEW RequestObject WITHOUT owner restriction to query across all owners
    ro_multi = RequestObject()
    ro_multi.set_http_method('GET')

    result_start = 0
    result_limit = RESULT_PAGE_SIZE

    while True:
        try:
            # NOTE: 
            ro_multi.set_request_uri(
                f'/v3/indicators?tql={tql_encoded}'
                f'&fields=tags,observations,associatedGroups,falsePositives,threatAssess'
                f'&resultStart={result_start}&resultLimit={result_limit}'
            )

            response = tc.api_request(ro_multi)

            ct = response.headers.get('content-type', '')
            if not ct.startswith('application/json'):
                raise RuntimeError(f"Non-JSON response ({ct}): {response.content[:200]}")

            results = response.json()
            data_items = results.get('data', []) or []

            # stop when no more results
            if not data_items:
                break

            final_results.append(results)
            result_start += result_limit

        except Exception as e:
            display(f"Failed to query indicators (start={result_start}): {e}")
            break

    # Normalize results
    normalized_data = []
    for result in final_results:
        data_items = result.get('data', [])
        if not data_items:
            display("No data returned in API response:", result)
        for item in data_items:
            if isinstance(item, dict) and 'summary' in item:
                normalized_data.append(item)

    if normalized_data:
        observed_src = pd.json_normalize(normalized_data)
        observed_src['indicator'] = observed_src['summary'].astype(str).str.split().str[0].str.strip()
        observed_src['lastObserved'] = pd.to_datetime(observed_src['lastObserved'], utc=True, errors='coerce')
        observed_src = observed_src[observed_src['lastObserved'] >= pd.to_datetime(start, utc=True)]

        # Create a 'sources' column by aggregating ownerName values per indicator
        sources_per_indicator = (
            observed_src.groupby('indicator')['ownerName']
            .apply(lambda x: ', '.join(sorted(set(x))))
            .reset_index()
            .rename(columns={'ownerName': 'sources'})
        )

        # Merge sources back into observed_src
        observed_src = observed_src.merge(sources_per_indicator, on='indicator', how='left')
        # Filter to keep only records where ownerName is 'HTOC Org'
        observed_src = observed_src[observed_src['ownerName'] == 'HTOC Org'].copy()
    else:
        display("No valid indicator data found.")
        observed_src = pd.DataFrame()

    display(observed_src)

    import pandas as pd

    # Configuration for observed tags
    TAGS_FILE_PATH = r"Z:\HTOC\Data_Analytics\Data\Observed_Tags\htoc_observed_indicator_tags.csv"
    THREAT_CATEGORY_FILTER = 'THREAT ACTOR'

    # Load the observed tags data from the CSV file
    tags_file_path = TAGS_FILE_PATH
    htoc_observed_tags = pd.read_csv(tags_file_path)

    # Filter for THREAT ACTOR category only
    threat_actor_records = htoc_observed_tags[
        htoc_observed_tags['threat_category'] == THREAT_CATEGORY_FILTER
    ]

    # Group by indicator and aggregate the tag information
    threat_actor_condensed = threat_actor_records.groupby('indicator').agg({
        'type': 'first',
        'orig_tag': lambda x: ', '.join(sorted(set(x.dropna()))),
        'tag': lambda x: ', '.join(sorted(set(x.dropna()))),
        'threat_category': lambda x: ', '.join(sorted(set(x.dropna()))),
        'NATION STATE': lambda x: ', '.join(sorted(set(x.dropna()))) if x.notna().any() else None,
        'SECURITY ORGANIZATION': lambda x: ', '.join(sorted(set(x.dropna()))) if x.notna().any() else None,
        'MALWARE CLASS': lambda x: ', '.join(sorted(set(x.dropna()))) if x.notna().any() else None,
        'CVE_NBR': lambda x: ', '.join(sorted(set(x.dropna()))) if x.notna().any() else None
    }).reset_index()

    print(f"Loaded {len(threat_actor_records)} {THREAT_CATEGORY_FILTER} rows from {TAGS_FILE_PATH}")
    print(f"Condensed to {len(threat_actor_condensed)} unique indicators")
    display(threat_actor_condensed)

    # Merge threat actor tags into observed_src
    threat_actor_tags = threat_actor_condensed[['indicator', 'tag']].copy()
    threat_actor_tags = threat_actor_tags.rename(columns={'tag': 'threat_actor'})
    observed_src = observed_src.merge(threat_actor_tags, on='indicator', how='left')
    print(f"Added threat_actor column to observed_src")
    print(f"observed_src columns after merge: {observed_src.columns.tolist()}")
    display(observed_src[['indicator', 'type', 'threat_actor']])

    import pandas as pd

    # Configuration for observed indicators file and first-seen window
    OBSERVED_INDICATORS_CSV_PATH = r"Z:\HTOC\Data_Analytics\Data\Observed_Indicators\htoc_observed_indicators.csv"
    FIRSTSEEN_LOOKBACK_DAYS = 360

    csv_path = OBSERVED_INDICATORS_CSV_PATH
    df_obs = pd.read_csv(csv_path)

    df_obs['firstseen_dt'] = pd.to_datetime(df_obs['firstseen_dt'], errors='coerce')

    today = pd.Timestamp.today().normalize()
    cutoff = today - pd.Timedelta(days=FIRSTSEEN_LOOKBACK_DAYS)

    df_last14 = df_obs[
        (df_obs['firstseen_dt'] >= cutoff) &
        (df_obs['firstseen_dt'] <= today)
    ][['indicator', 'firstseen_dt']]

    df_last14 = df_last14.rename(columns={'firstseen_dt': 'firstseen_date'})

    observed_src = observed_src.merge(
        df_last14[['indicator', 'firstseen_date']],
        on='indicator',
        how='left'
    )

    observed_src

    # Add incidents/events as columns on observed_src

    INCIDENT_COLUMN_NAME = "incidents/events"
    GROUP_TYPES_OF_INTEREST = {"incident", "event"}

    import re
    INCIDENT_ID_REGEX = re.compile(r"\bINC\d+\b", re.IGNORECASE)

    if observed_src.empty:

        observed_src[INCIDENT_COLUMN_NAME] = []

        incidents_df = pd.DataFrame()

        display("No observed data available to annotate with incidents/events.")

    elif 'associatedGroups.data' not in observed_src.columns:

        observed_src[INCIDENT_COLUMN_NAME] = "None"

        incidents_df = pd.DataFrame()

        display("associatedGroups.data not present; created placeholder incidents/events column with 'None'.")

    else:

        def _extract_groups_of_interest(val):

            if isinstance(val, list):

                return [item for item in val if isinstance(item, dict) and str(item.get('type', '')).lower() in GROUP_TYPES_OF_INTEREST]

            if isinstance(val, dict):

                return [val] if str(val.get('type', '')).lower() in GROUP_TYPES_OF_INTEREST else []

            return []

        def _format_group_label(item):

            if not isinstance(item, dict):

                return str(item)

            typ = str(item.get('type', '')).title()

            number = item.get('id') or item.get('xid') or item.get('name')

            return f"{typ}:{number}" if number is not None else typ

        def _extract_incidents_from_description(text):

            if not isinstance(text, str):

                return []

            matches = INCIDENT_ID_REGEX.findall(text)

            # Preserve order but remove duplicates and normalize to upper
            unique_ids = list(dict.fromkeys(m.upper() for m in matches))

            return [
                {"type": "incident", "id": inc_id}
                for inc_id in unique_ids
            ]

        groups = observed_src['associatedGroups.data'].apply(_extract_groups_of_interest)

        if 'description' in observed_src.columns:

            desc_groups = observed_src['description'].apply(_extract_incidents_from_description)

        else:

            desc_groups = pd.Series([[] for _ in range(len(observed_src))], index=observed_src.index)

        combined_groups = groups.combine(desc_groups, lambda a, b: (a or []) + (b or []))

        observed_src[INCIDENT_COLUMN_NAME] = combined_groups.apply(
            lambda lst: ";".join(_format_group_label(it) for it in lst) if isinstance(lst, list) and len(lst) > 0 else "None"
        )

        incidents_df = observed_src[combined_groups.apply(lambda lst: isinstance(lst, list) and len(lst) > 0)].copy()

        display(f"Annotated incidents/events for {len(observed_src)} indicators; found {len(incidents_df)} with matches.")

        display_cols = ['indicator', 'type', INCIDENT_COLUMN_NAME]

        display(incidents_df[display_cols])

    # Unnest tags.data in observed_src to get a list of each tag per indicator

    # Explode tags.data to one row per tag
    tags_exploded = (
        observed_src[['indicator', 'tags.data']]
          .explode('tags.data')
          .dropna(subset=['tags.data'])
    )

    # Extract tag name from each tag object
    tags_exploded['tag_name'] = tags_exploded['tags.data'].apply(lambda x: x.get('name') if isinstance(x, dict) else None)

    # Aggregate all tag names into a list per indicator
    tags_per_indicator = (
        tags_exploded.groupby('indicator')['tag_name']
        .apply(lambda x: [t for t in x if t])
        .reset_index()
        .rename(columns={'tag_name': 'tag_list'})
    )

    # Merge back to observed_src
    observed_src_with_tags = observed_src.merge(tags_per_indicator, on='indicator', how='left')

    # Optional: show indicators with 'Scan' in tag_list
    # display(observed_src_with_tags[observed_src_with_tags['tag_list'].apply(lambda tags: isinstance(tags, list) and any(isinstance(tag, str) and 'Scan' in tag for tag in tags))])

    # --- Botnet tags: count and add Botnet column to observed_src ---
    BOTNET_TAGS_OF_INTEREST = [
        "Scanning", "DDoS", "Spam", "Phishing", "Cryptojacking",
        "Credential Stuffing", "Ransomware", "Data Theft",
        "Cross Site Scripting Attacks", "SQL Injections",
    ]
    tags_of_interest = BOTNET_TAGS_OF_INTEREST

    def tag_counter(tag_list, tag):
        if not isinstance(tag_list, list):
            return 0
        tag_lc = tag.lower()
        return sum(1 for t in tag_list if isinstance(t, str) and t.strip().lower() == tag_lc)

    tag_counts = {}
    for tag in tags_of_interest:
        tag_counts[tag] = observed_src_with_tags['tag_list'].apply(lambda lst: tag_counter(lst, tag)).sum()

    botnet_tags = set(tags_of_interest)
    botnet_tags_lower = {b.lower() for b in botnet_tags}

    def extract_botnet_tags(tag_list):
        if not isinstance(tag_list, list):
            return []
        return [t for t in tag_list if isinstance(t, str) and t.strip().lower() in botnet_tags_lower]

    indicator_to_tags = dict(zip(observed_src_with_tags['indicator'], observed_src_with_tags['tag_list']))
    observed_src['Botnet'] = observed_src['indicator'].map(lambda ind: extract_botnet_tags(indicator_to_tags.get(ind, [])))

    pd.DataFrame(list(tag_counts.items()), columns=['Tag', 'Count'])


    import os
    import pandas as pd
    from datetime import datetime, timedelta

    # Configuration for OpDiv observation files
    OPDIV_BASE_PATH = r"Z:/HTOC/Data_Analytics/Data/OpDiv_Observations/htoc_opdiv_obs_d{date}.csv"
    #OPDIV_BASE_PATH = r"C:\Users\jaskew\Documents\project_repository\data\raw\ObservationDataFiles\htoc_opdiv_obs_d{date}.csv"
    OPDIV_DATE_FORMAT = "%Y%m%d"
    OPDIV_LOOKBACK_DAYS = 360

    # Base file path with placeholder for date
    base_path = OPDIV_BASE_PATH

    def get_file_paths(base_path, days=7):
        """Generate file paths for the last `days` days using list comprehension."""
        today = datetime.utcnow()
        dates_to_pull = [(today - timedelta(days=i)).strftime(OPDIV_DATE_FORMAT) for i in range(days)]

        # Construct file paths
        file_paths = [base_path.format(date=dt) for dt in dates_to_pull]

        # Filter for existing files
        existing_files = [file_path for file_path in file_paths if os.path.exists(file_path)]

        if not existing_files:
            display("No files found for the specified date range.")
        else:
            display(f"Files to be loaded: {existing_files}")

        return existing_files

    def load_observed_data(file_paths):
        """Load and concatenate observed data from multiple files."""
        data_frames = []

        for file_path in file_paths:
            try:
                df = pd.read_csv(file_path)
                data_frames.append(df)
            except Exception as e:
                display(f"Error reading file {file_path}: {e}")

        # Concatenate data
        if data_frames:
            observed_data_df = pd.concat(data_frames, ignore_index=True)
            display(f"Loaded data from {len(data_frames)} files.")
        else:
            observed_data_df = pd.DataFrame()

        return observed_data_df

    # Fetch file paths for the last OPDIV_LOOKBACK_DAYS days
    file_paths = get_file_paths(base_path, days=OPDIV_LOOKBACK_DAYS)

    # Load observed data
    observed_data_df = load_observed_data(file_paths)

    # ── Mass Scanner Detection (Tiered) ───────────────────────────────────────────
    # Tier 1: 10K–100K obs in 7 days + 5+ OpDivs — moderate raw-score multiplier (no PRISM cap).
    # Tier 2: 100K+ obs in 7 days + 5+ OpDivs — heavy raw-score multiplier (no PRISM cap).
    # Strong context (threat actor, incidents, TOR, etc.) can still raise the final score.
    MASS_SCANNER_TIER1_OBS  = 10_000   # Tier 1 lower bound
    MASS_SCANNER_TIER2_OBS  = 100_000  # Tier 2 threshold
    MASS_SCANNER_OPDIV_MIN  = 5        # must hit at least 5 OpDivs (broad spread)

    if not observed_data_df.empty and 'obs_date' in observed_data_df.columns:
        cutoff_7d = (pd.Timestamp.utcnow() - pd.Timedelta(days=7)).strftime('%Y-%m-%d')
        obs_7d = observed_data_df[observed_data_df['obs_date'] >= cutoff_7d].copy()

        mass_scanner_agg = (
            obs_7d.groupby('indicator')
            .agg(
                total_obs_7d    =('observations', 'sum'),
                unique_opdivs_7d=('OpDiv', 'nunique'),
                unique_days_7d  =('obs_date', 'nunique'),
            )
            .reset_index()
        )

        broad_spread = mass_scanner_agg['unique_opdivs_7d'] >= MASS_SCANNER_OPDIV_MIN

        mass_scanner_agg['mass_scanner_tier1'] = (
            (mass_scanner_agg['total_obs_7d'] >= MASS_SCANNER_TIER1_OBS) &
            (mass_scanner_agg['total_obs_7d'] <  MASS_SCANNER_TIER2_OBS) &
            broad_spread
        )
        mass_scanner_agg['mass_scanner_tier2'] = (
            (mass_scanner_agg['total_obs_7d'] >= MASS_SCANNER_TIER2_OBS) &
            broad_spread
        )

        n1 = mass_scanner_agg['mass_scanner_tier1'].sum()
        n2 = mass_scanner_agg['mass_scanner_tier2'].sum()
        display(f"Mass scanner Tier 1 (moderate penalty): {n1} indicators "
                f"({MASS_SCANNER_TIER1_OBS:,}–{MASS_SCANNER_TIER2_OBS:,} obs, {MASS_SCANNER_OPDIV_MIN}+ OpDivs)")
        display(f"Mass scanner Tier 2 (heavy penalty):  {n2} indicators "
                f"(>= {MASS_SCANNER_TIER2_OBS:,} obs, {MASS_SCANNER_OPDIV_MIN}+ OpDivs)")
        display(mass_scanner_agg[mass_scanner_agg['mass_scanner_tier2']][
            ['indicator', 'total_obs_7d', 'unique_opdivs_7d', 'unique_days_7d']
        ].sort_values('total_obs_7d', ascending=False))
    else:
        mass_scanner_agg = pd.DataFrame(columns=[
            'indicator', 'total_obs_7d', 'unique_opdivs_7d', 'unique_days_7d',
            'mass_scanner_tier1', 'mass_scanner_tier2'
        ])
        display("Mass scanner detection skipped — observed_data_df is empty or missing obs_date.")

    import pandas as pd
    from datetime import timedelta
    import warnings

    warnings.simplefilter(action='ignore', category=pd.errors.SettingWithCopyWarning)

    # ═══════════════════════════════════════════════════════════════════════════════
    # CONFIGURATION & SETUP
    # ═══════════════════════════════════════════════════════════════════════════════

    # Time cutoffs
    cutoff = pd.Timestamp.utcnow()
    cutoff_naive = cutoff.tz_convert(None)

    # Define known partner names (standalone tags that represent partners)
    KNOWN_PARTNERS = {'DHA', 'OS', 'FDA', 'CMS', 'VA', 'HRSA', 'NIH', 'IHS', 'HHS', 'CDC'}

    # ═══════════════════════════════════════════════════════════════════════════════
    # HELPER FUNCTIONS
    # ═══════════════════════════════════════════════════════════════════════════════

    def get_all_partner_indicators_from_obs(observed_data_df, cutoff_naive):
        """Get ALL indicators that have partners from observed_data_df (no minimum threshold)."""
        if observed_data_df.empty or 'OpDiv' not in observed_data_df.columns:
            return pd.DataFrame()

        # Ensure obs_date is datetime
        observed_data_df['obs_date'] = pd.to_datetime(observed_data_df['obs_date'], errors='coerce')

        # Filter by recent dates (last 60 days to match wider time window)
        recent_obs = observed_data_df[
            observed_data_df['obs_date'] >= cutoff_naive - timedelta(days=60)
        ].copy()

        if recent_obs.empty:
            return pd.DataFrame()

        # Group by indicator and count unique OpDiv (partners)
        partner_counts = (
            recent_obs.groupby('indicator')['OpDiv']
            .agg(['nunique', lambda s: ', '.join(sorted(set(s.dropna())))]).reset_index()
            .rename(columns={'nunique': 'partner_count_obs', '<lambda_0>': 'partners_from_obs'})
        )

        # Keep ALL indicators with partners (no minimum threshold)
        all_partner_indicators = partner_counts[partner_counts['partner_count_obs'] >= 1].copy()

        return all_partner_indicators

    def extract_partners_from_tags(observed_src):
        """Extract partner information from ThreatConnect tags."""
        df = observed_src.copy()

        # explode tags.data
        tags_exploded = (
            df[['indicator', 'tags.data']]
              .explode('tags.data')
              .dropna(subset=['tags.data'])
        )

        # normalize all fields in tags.data into flat columns
        tags_norm = pd.json_normalize(tags_exploded['tags.data'])
        tags_norm.columns = [f"tag_{c}" for c in tags_norm.columns]

        # Replace VA CSOC CTS Splunk with VA Splunk API in tag_name
        tags_norm['tag_name'] = tags_norm['tag_name'].str.replace('VA CSOC CTS Splunk', 'VA Splunk API', regex=False)

        # re‑attach indicator
        tags_expanded = tags_exploded.reset_index(drop=True).join(tags_norm)

        # extract partners from tags ending with " Splunk API"
        tags_expanded['partner'] = tags_expanded['tag_name'].map(
            lambda n: n[:-len(' Splunk API')] if isinstance(n, str) and n.endswith(' Splunk API') else None
        )

        # aggregate each tag_* field into a list per indicator
        tag_fields = list(tags_norm.columns)
        tag_agg = (
            tags_expanded
              .groupby('indicator', as_index=False)
              .agg({
                  **{f: list for f in tag_fields},
                  'partner': lambda x: [p for p in dict.fromkeys(x) if p]
              })
              .rename(columns={'partner':'partners_from_tags'})
        )

        # Also extract partners from standalone tags in tag_list (tag_name column)
        def extract_standalone_partners(tag_list):
            """Extract partner names from tag_list that match known partners"""
            if not isinstance(tag_list, list):
                return []
            # Check each tag in the list to see if it matches a known partner name
            found_partners = []
            for tag in tag_list:
                if isinstance(tag, str) and tag.strip() in KNOWN_PARTNERS:
                    found_partners.append(tag.strip())
            return found_partners

        # Extract standalone partners from tag_name (which contains the tag_list)
        if 'tag_name' in tag_agg.columns:
            tag_agg['standalone_partners'] = tag_agg['tag_name'].apply(extract_standalone_partners)

            # Combine partners from both sources (Splunk API tags and standalone tags)
            def combine_tag_partners(row):
                """Combine partners from Splunk API tags and standalone tags"""
                partners_from_splunk = row.get('partners_from_tags', [])
                partners_standalone = row.get('standalone_partners', [])

                # Handle both list and string formats
                if isinstance(partners_from_splunk, str):
                    partners_from_splunk = [p.strip() for p in partners_from_splunk.split(',') if p.strip()]
                if not isinstance(partners_from_splunk, list):
                    partners_from_splunk = []

                # Combine and deduplicate
                all_partners = list(dict.fromkeys(partners_from_splunk + partners_standalone))
                return all_partners

            tag_agg['partners_from_tags'] = tag_agg.apply(combine_tag_partners, axis=1)

            # Drop the temporary standalone_partners column
            tag_agg = tag_agg.drop(columns=['standalone_partners'], errors='ignore')

        return tag_agg, tag_fields

    def combine_partners_from_sources(base_agg, tag_agg, all_partner_indicators):
        """Combine partner information from both observation data and tags."""
        # Merge tag aggregation
        agg_df = base_agg.merge(tag_agg, on='indicator', how='left')

        # Merge with all_partner_indicators to get partners from observed_data_df
        if not all_partner_indicators.empty:
            agg_df = agg_df.merge(
                all_partner_indicators[['indicator', 'partners_from_obs', 'partner_count_obs']],
                on='indicator',
                how='left'
            )
        else:
            # Add empty columns if no partner indicators found
            agg_df['partners_from_obs'] = ''
            agg_df['partner_count_obs'] = 0

        # Combine partners from all sources
        def combine_all_partners(row):
            """Combine partners from observed_data_df and ThreatConnect tags."""
            obs_partners = row.get('partners_from_obs', '')
            tag_partners = row.get('partners_from_tags', [])

            combined = set()

            # Add partners from observed_data_df
            if pd.notna(obs_partners) and obs_partners:
                for p in str(obs_partners).split(', '):
                    if p.strip():
                        combined.add(p.strip())

            # Add partners from tags
            if isinstance(tag_partners, list):
                for p in tag_partners:
                    if p and p.strip():
                        combined.add(p.strip())
            elif pd.notna(tag_partners) and tag_partners:
                # Handle case where tag_partners might be a string
                for p in str(tag_partners).split(','):
                    if p.strip():
                        combined.add(p.strip())

            return ', '.join(sorted(combined)) if combined else ''

        # Create combined partners column
        agg_df['partners'] = agg_df.apply(combine_all_partners, axis=1)

        # Calculate partner count
        agg_df['partner_count'] = agg_df['partners'].apply(
            lambda x: len([p for p in str(x).split(', ') if p.strip()]) if pd.notna(x) and x else 0
        )

        # Clean up temporary columns
        cols_to_drop = ['partners_from_obs', 'partner_count_obs', 'partners_from_tags']
        agg_df = agg_df.drop(columns=[col for col in cols_to_drop if col in agg_df.columns], errors='ignore')

        return agg_df

    # ═══════════════════════════════════════════════════════════════════════════════
    # MAIN PROCESSING PIPELINE
    # ═══════════════════════════════════════════════════════════════════════════════

    print("Starting enhanced partner extraction pipeline...")

    # Step 1: Get ALL indicators with partners from observed_data_df (no filtering)
    print("Identifying ALL indicators with partners from observed_data_df...")
    all_partner_indicators = get_all_partner_indicators_from_obs(observed_data_df, cutoff_naive)
    print(f"Found {len(all_partner_indicators)} indicators with partners from observation data")

    # Step 2: Extract partners from ThreatConnect tags
    print("Extracting partners from ThreatConnect tags...")
    tag_agg, tag_fields = extract_partners_from_tags(observed_src)
    print(f"Processed tags for {len(tag_agg)} indicators")

    # Step 3: Core aggregation of other columns
    print("Performing core aggregation...")
    df = observed_src.copy()

    first_cols = [
        'id','dateAdded','ownerId','ownerName','webLink','type','lastModified', 'falsePositives',
        'rating','confidence','description','summary','observations',
        'lastObserved','privateFlag','active','activeLocked','ip',
        'legacyLink','source','address','url', 'threatAssessScore', 'calScore', 'incidents/events', 'sources',
        'threat_actor','firstseen_date'
    ]

    # Add 'Botnet' column from observed_src if it exists
    if 'Botnet' in observed_src.columns:
        df['Botnet'] = observed_src['Botnet']
        first_cols.append('Botnet')

    base_agg = (
        df
          .drop(columns=[
              'createdBy.id','createdBy.userName','createdBy.firstName','createdBy.lastName',
              'createdBy.pseudonym','createdBy.owner','xid','eventType','documentDateAdded',
              'documentType','fileSize','fileName','downVoteCount','upVoteCount','type_group',
              'webLink_group','ownerName_group','ownerId_group','dateAdded_group','id_group',
              'platforms.count','tactics.count',
          ], errors='ignore')
          .groupby('indicator', as_index=False)[
              [c for c in first_cols if c in df.columns]
          ]
          .first()
    )

    # Step 4: Combine partners from all sources
    print("Combining partners from all sources...")
    agg_df = combine_partners_from_sources(base_agg, tag_agg, all_partner_indicators)

    # Step 5: Clean up and format list columns
    print("Cleaning up list columns...")
    def clean_list(lst):
        if not isinstance(lst, list):
            return []
        cleaned = []
        for v in lst:
            # drop any null‑like values
            try:
                if pd.isna(v):
                    continue
            except Exception:
                pass
            # drop empty strings
            if isinstance(v, str) and v == "":
                continue
            cleaned.append(v)
        return cleaned

    def list_to_csv(lst):
        """
        Takes a cleaned list and returns:
          - a comma-separated string of its items, OR
          - an empty string if there are none.
        """
        if not lst:
            return ""
        return ", ".join(str(v) for v in lst)

    # Apply cleaning to tag fields (but not partners, which is already a string)
    for col in ['group_ids', 'group_names'] + tag_fields:
        if col in agg_df.columns:
            agg_df[col] = agg_df[col].apply(clean_list).apply(list_to_csv)

    print(f"Processing complete! Final dataset has {len(agg_df)} indicators")

    # Merge mass scanner tier flags computed in Cell 8
    if not mass_scanner_agg.empty:
        agg_df = agg_df.merge(
            mass_scanner_agg[[
                'indicator', 'total_obs_7d', 'unique_opdivs_7d',
                'mass_scanner_tier1', 'mass_scanner_tier2'
            ]],
            on='indicator',
            how='left'
        )
        agg_df['mass_scanner_tier1'] = agg_df['mass_scanner_tier1'].fillna(False).astype(bool)
        agg_df['mass_scanner_tier2'] = agg_df['mass_scanner_tier2'].fillna(False).astype(bool)
        agg_df['total_obs_7d']       = agg_df['total_obs_7d'].fillna(0).astype(int)
        agg_df['unique_opdivs_7d']   = agg_df['unique_opdivs_7d'].fillna(0).astype(int)
        print(f"Mass scanner flags merged — Tier 1: {agg_df['mass_scanner_tier1'].sum()} | Tier 2: {agg_df['mass_scanner_tier2'].sum()}")
    else:
        agg_df['mass_scanner_tier1'] = False
        agg_df['mass_scanner_tier2'] = False
        agg_df['total_obs_7d']       = 0
        agg_df['unique_opdivs_7d']   = 0

    display(agg_df)
    # ──────────────────────────────────────────────────────────────────────────────
    # Clean enrichment: filter unsupported types, use ID when available, parallelize
    # ──────────────────────────────────────────────────────────────────────────────
    import urllib.parse
    import pandas as pd
    from concurrent.futures import ThreadPoolExecutor, as_completed

    COL_PATH = "data.enrichment.data"   # adjust if API changes

    # Determine candidate indicators (use 'indicator' if present, else 'summary')
    key_col = 'indicator' if 'indicator' in agg_df.columns else 'summary'

    # Provider allowlists: only call API for types that support VT and/or Shodan
    VT_TYPES = {'Address', 'IPv4', 'IPv6', 'Host', 'Domain', 'URL', 'File', 'SHA1', 'SHA256', 'MD5'}
    SHODAN_TYPES = {'Address', 'IPv4', 'IPv6'}

    # Build candidates and include 'id' when present (for ID-based lookup)
    cols = [key_col, 'type']
    if 'id' in agg_df.columns:
        cols.append('id')

    candidates = (
        agg_df[cols]
        .dropna(subset=[key_col])
        .astype({key_col: str})
        .drop_duplicates(subset=[key_col])
    )

    # Pre-filter: skip types that cannot be enriched (e.g. EmailAddress)
    candidates = candidates[
        candidates['type'].astype(str).str.strip().isin(VT_TYPES | SHODAN_TYPES)
    ].copy()

    indicator_values = candidates[key_col].tolist()
    display(f"Enriching {len(indicator_values)} indicators (VT; Shodan for IP types only)...")

    def _enrich_one(row_series):
        """Call enrich API for one indicator. Returns (data_dict, None) on success or (None, fail_dict) on failure."""
        value = row_series[key_col]
        typ = str(row_series.get('type', '') or '')
        row_id = row_series.get('id')
        use_id = pd.notna(row_id) and str(row_id).strip().isdigit()

        try:
            # Prefer ID-based URL to avoid "More than one indicator matches" errors
            if use_id:
                indicator_id_or_value = str(int(float(row_id)))
            else:
                indicator_id_or_value = urllib.parse.quote(value, safe="")

            providers = []
            if typ in VT_TYPES:
                providers.append("VirusTotalV3")
            if typ in SHODAN_TYPES:
                providers.append("Shodan")
            if not providers:
                providers.append("VirusTotalV3")

            q = urllib.parse.urlencode({"type": providers}, doseq=True)
            enrich_url = f"/v3/indicators/{indicator_id_or_value}/enrich?{q}"

            ro = RequestObject()
            ro.set_http_method("POST")
            ro.set_request_uri(enrich_url)
            ro.set_body({})
            resp = tc.api_request(ro)

            try:
                data = resp.json()
            except Exception:
                data = {"status": getattr(resp, 'status_code', 'n/a'), "raw": getattr(resp, 'text', None)}

            data[key_col] = value
            return (data, None)
        except Exception as e:
            return (None, {key_col: value, "type": typ, "error": str(e)})

    enriched_results = []
    failed = []
    max_workers = 8

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(_enrich_one, row): row for _, row in candidates.iterrows()}
        for future in as_completed(futures):
            result, err = future.result()
            if result is not None:
                enriched_results.append(result)
            else:
                failed.append(err)

    # If nothing enriched, carry on with base
    if not enriched_results:
        display("No enrichment data retrieved.")
        recent_tags = agg_df.copy()
    else:
        df_enriched = (
            pd.json_normalize(enriched_results)
            .drop_duplicates(subset=[key_col], keep="last")
        )

        # Merge enrichment block into base
        recent_tags = agg_df.merge(df_enriched, on=key_col, how="left")

        # ── Flatten enrichment payload without creating duplicate base rows ───────
        if COL_PATH in recent_tags.columns:
            exploded = (
                recent_tags[[key_col, COL_PATH]]
                .explode(COL_PATH)
                .dropna(subset=[COL_PATH])
            )

            enrich_flat = pd.json_normalize(exploded[COL_PATH]).add_prefix("enrich_")
            enrich_flat[key_col] = exploded[key_col].values

            def _flatten_lists(values):
                out = []
                for v in values:
                    if isinstance(v, list):
                        out.extend(v)
                    else:
                        out.append(v)
                return out

            def _agg_obj(series: pd.Series):
                vals = [v for v in series.dropna()]
                if not vals:
                    return None
                flat = _flatten_lists(vals)
                if all(not isinstance(v, (list, dict)) for v in flat):
                    arr = pd.Series(flat).unique()
                    return list(arr)
                return flat

            obj_cols = enrich_flat.select_dtypes("object").columns.difference([key_col])
            num_cols = enrich_flat.columns.difference(obj_cols.union({key_col}))

            agg_dict = {c: _agg_obj for c in obj_cols}
            agg_dict.update({c: "max" for c in num_cols})

            enrich_wide = (
                enrich_flat
                .groupby(key_col, as_index=False)
                .agg(agg_dict)
            )

            recent_tags = (
                recent_tags.drop(columns=[COL_PATH], errors="ignore")
                           .drop_duplicates(subset=[key_col])
                           .merge(enrich_wide, on=key_col, how="left")
            )

        display(f"Enrichment complete for {recent_tags[key_col].notna().sum()} indicators.")

    # Compact failure summary without flooding output
    if failed:
        fail_df = pd.DataFrame(failed)
        display(f"{len(failed)} indicators failed enrichment (showing up to 10):")
        display(fail_df.head(10))

    # Show preview of key enrichment columns if present
    preview_cols = [c for c in [key_col, 'enrich_hostNames', 'enrich_domains', 'enrich_tags', 'enrich_vtMaliciousCount'] if c in recent_tags.columns]

    recent_tags.drop(
        columns=[
            'tag_id', 'tag_lastUsed', 'tag_lastModified', 'tag_ownerId',
            'tag_ownerName', 'tag_dateAdded', 'tag_description','tag_tactics.count',
            'tag_platform.data', 'tag_platform.count', 'data.id', 'data.dateAdded', 'data.ownerId',
            'data.webLink', 'data.ownerName', 'data.lastModified', 'data.summary', 'data.ip',
            'data.legacyLink','data.source', 'enrich_cloudProvider', 'enrich_cloudRegion', 'enrich_type',
            'id'
        ],
        inplace=True,
        errors='ignore'
    )

    display(recent_tags[preview_cols])


    # In[28]:


    # Count how many indicators are associated with a unique enrich_domains value

    # Only proceed if 'enrich_domains' exists in exploded DataFrame
    if 'enrich_domains' in exploded.columns:
        # Drop rows where enrich_domains is missing or NaN
        domains_df = exploded.dropna(subset=['enrich_domains'])

        # If enrich_domains is a list, flatten to individual domain rows
        def flatten_domains(row):
            val = row['enrich_domains']
            if isinstance(val, list):
                return [(row['indicator'], d) for d in val if pd.notna(d)]
            elif pd.notna(val):
                return [(row['indicator'], val)]
            return []

        flat = []
        for _, row in domains_df.iterrows():
            flat.extend(flatten_domains(row))

        flat_df = pd.DataFrame(flat, columns=['indicator', 'domain'])

        # Count unique indicators per domain
        domain_counts = (
            flat_df.groupby('domain')['indicator']
            .nunique()
            .reset_index()
            .rename(columns={'indicator': 'indicator_count'})
            .sort_values('indicator_count', ascending=False)
        )

        display(domain_counts)
    else:
        display("Column 'enrich_domains' does not exist in the provided DataFrame. Skipping domain association count.")

    # Re-use get_file_paths/load_observed_data to pull a longer window (365 days), then add obs_count to recent_tags
    OPDIV_OBS_COUNT_DAYS = 365
    file_paths = get_file_paths(base_path, days=OPDIV_OBS_COUNT_DAYS)
    observed_data_df = load_observed_data(file_paths)

    # Aggregate by indicator (unique obs_date count) and merge into recent_tags
    agg_by_indicator = (
        observed_data_df
        .groupby('indicator', as_index=False)['obs_date']
        .nunique()
        .rename(columns={'obs_date': 'obs_days_count'})
    )
    agg_by_indicator = agg_by_indicator[agg_by_indicator['indicator'].isin(recent_tags['indicator'])]
    recent_tags = recent_tags.merge(
        agg_by_indicator.rename(columns={'obs_days_count': 'obs_count'}),
        on='indicator',
        how='left'
    )
    display(agg_by_indicator)

    import pandas as pd
    import numpy as np

    # -------------------------------------------------------------------
    # 0. Load your data (replace this with your actual load)
    # -------------------------------------------------------------------
    # recent_tags = pd.read_csv("recent_tags.csv")

    df_scored = recent_tags.copy()

    # -------------------------------------------------------------------
    # 0.5 Missing-value modification (VT presence + display label)
    #   - VT missing is NEUTRAL: no penalty, no boost.
    # -------------------------------------------------------------------
    VT_COL = 'enrich_vtMaliciousCount'
    # Scoring ceiling: indicators at or above this are treated as fully malicious.
    # 94 is the theoretical max but malicious intent is clear well before that.
    VT_EFFECTIVE_MAX = 40

    if VT_COL in df_scored.columns:
        df_scored[VT_COL] = pd.to_numeric(df_scored[VT_COL], errors='coerce')
        df_scored['vt_present'] = df_scored[VT_COL].notna()
    else:
        df_scored[VT_COL] = np.nan
        df_scored['vt_present'] = False

    df_scored['vt_present'] = df_scored['vt_present'].astype(bool)

    # Display-friendly VT column (shows raw value, not capped)
    df_scored['vt_display'] = np.where(df_scored['vt_present'], df_scored[VT_COL], 'No VT Score')

    # Numeric-safe VT for math (missing -> 0; capped at effective max for scoring)
    df_scored['vt_numeric_for_scoring'] = df_scored[VT_COL].fillna(0).clip(0, VT_EFFECTIVE_MAX)

    # -------------------------------------------------------------------
    # 1. Rule-based Threat Scoring
    # -------------------------------------------------------------------

    # ── Severity Binning ────────────────────────────
    scoring_bins = [0, 200, 500, 800, 1001]
    label_bins = ['low', 'medium', 'high', 'critical']

    # ── Feature Weights / Params ────────────────────
    Weights = {
        "MALICIOUS_WEIGHT": 7.50,
        "OBSERVATION_COUNT_WEIGHT": 0.02,
        "CONTINUITY_WEIGHT": 0.90,
        "TC_RATING": 0.01,
        "TC_CONFIDENCE": 0.025,
        "TOR_ACTIVITY_WEIGHT": 9.00,
        "CAL_SCORE_WEIGHT": 2.75,
        "TC_THREAT_SCORE_WEIGHT": 0.75,
        "INCIDENTS_EVENTS_WEIGHT": 8.00,
        "PARTNER_WEIGHT": 2.10,
        "SOURCES_WEIGHT": 2.80,
        "THREAT_ACTOR_WEIGHT": 10.00,
        "FIRST_OBS_WEIGHT": 2.00
    }

    # ── First-seen recency boost (additive, decaying) ──────────────────
    FIRST_OBS_MAX_DAYS = 14  # days back where boost fades to 0

    firstseen_col = df_scored.get('firstseen_date', pd.Series(pd.NaT, index=df_scored.index))
    firstseen_dt = pd.to_datetime(firstseen_col, errors='coerce')

    today = pd.Timestamp.today().normalize()
    age_days = (today - firstseen_dt).dt.days
    age_days = age_days.clip(lower=0)

    # Linear decay: 1.0 when age=0, 0.0 when age>=FIRST_OBS_MAX_DAYS
    freshness = (FIRST_OBS_MAX_DAYS - age_days) / FIRST_OBS_MAX_DAYS
    freshness = freshness.clip(lower=0.0, upper=1.0)

    # Neutral when no firstseen_date
    freshness = freshness.where(firstseen_dt.notna(), 0.0)

    df_scored['first_obs_raw_score'] = freshness * Weights['FIRST_OBS_WEIGHT']

    BOTNET_ACTIONS = {
        'scanning', 'ddos', 'spam', 'phishing', 'cryptojacking',
        'credential stuffing', 'ransomware'
    }

    TOR_ACTIVITY = {'tor', 'tor activity'}

    # ── Utility Functions ────────────────────────────────────────────────
    def convert_to_list(val):
        """
        Convert various input types to a list format.
        Handles: None, NaN, lists, tuples, sets, strings (including list representations and comma-separated).
        """
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return []
        if isinstance(val, (list, set, tuple)):
            return list(val)
        if isinstance(val, str):
            if val.strip().startswith('[') and val.strip().endswith(']'):
                try:
                    import ast
                    parsed = ast.literal_eval(val)
                    if isinstance(parsed, (list, tuple)):
                        return list(parsed)
                except Exception:
                    pass
            return [x.strip() for x in val.split(',') if x.strip()]
        return [val] if val else []

    # Known feature maximums (absolute, domain-informed)
    MAX_OBS_REALISTIC = 365
    MAX_RATING = 5
    MAX_CONFIDENCE = 100

    # Policy multipliers
    FALSE_POSITIVE_WEIGHT = 0.9          # 10% penalty
    BOTNET_MULTIPLIER     = 0.4          # 60% penalty when botnet-related
    SCANNER_PENALTY_MULTIPLIER = 0.80    # 20% penalty for scanner/benign-crawler tags
    DATA_QUALITY_FLOOR    = 0.85         # at worst 15% reduction for poor completeness

    # ── Input caps ──────────────────────────────────
    df_scored['obs_count'] = pd.to_numeric(
        df_scored['obs_count'] if 'obs_count' in df_scored.columns else pd.Series(0, index=df_scored.index),
        errors='coerce'
    ).fillna(0).clip(0, MAX_OBS_REALISTIC)

    df_scored['rating'] = pd.to_numeric(
        df_scored['rating'] if 'rating' in df_scored.columns else pd.Series(0, index=df_scored.index),
        errors='coerce'
    ).fillna(0).clip(0, MAX_RATING)

    df_scored['confidence'] = pd.to_numeric(
        df_scored['confidence'] if 'confidence' in df_scored.columns else pd.Series(0, index=df_scored.index),
        errors='coerce'
    ).fillna(0).clip(0, MAX_CONFIDENCE)

    df_scored['calScore'] = pd.to_numeric(
        df_scored['calScore'] if 'calScore' in df_scored.columns else pd.Series(0, index=df_scored.index),
        errors='coerce'
    ).fillna(0).clip(0, 1000)

    df_scored['type'] = (
        df_scored['type'] if 'type' in df_scored.columns else pd.Series('', index=df_scored.index)
    ).astype(str)

    # -------------------------------------------------------------------
    # VT Missing Behavior: NEUTRAL
    # -------------------------------------------------------------------
    df_scored['w_malicious_eff'] = Weights['MALICIOUS_WEIGHT']
    df_scored['w_tc_rating_eff'] = Weights['TC_RATING']

    # ── Base additive evidence ───────────────
    # Power scaling: compresses the VT range so mid-range scores still contribute
    # meaningfully without a linear jump at the top end.
    MALICIOUS_EXPONENT = 0.75
    df_scored['malicious_scaled'] = np.power(df_scored['vt_numeric_for_scoring'], MALICIOUS_EXPONENT)
    df_scored['malicious_raw_score'] = df_scored['malicious_scaled'] * Weights['MALICIOUS_WEIGHT']

    # ── FILE TYPE DETECTION ───────────────────────────
    FILE_TYPES = {'SHA1', 'SHA256', 'MD5', 'file', 'File'}
    df_scored['is_file_type'] = df_scored['type'].isin(FILE_TYPES)

    df_scored['continuity_val'] = df_scored['type'].map({
        'Address': 1, 'IPv4': 1, 'IPv6': 1,
        'Domain': 2, 'Host': 2, 'URL': 2, 'Stripped URL': 2,
        'EmailAddress': 2, 'EmailSubject': 2,
        'SHA1': 3, 'SHA256': 3, 'MD5': 3, 'file': 3, 'File': 3
    }).fillna(0)

    # For file types, override continuity score with fixed 900-point base
    df_scored['continuity_raw_score'] = df_scored['continuity_val'] * Weights['CONTINUITY_WEIGHT']
    df_scored.loc[df_scored['is_file_type'], 'continuity_raw_score'] = 900

    df_scored['tc_raw_rating'] = df_scored['rating'] * df_scored['w_tc_rating_eff']

    # Compress TC confidence so it contributes without dominating
    df_scored['tc_raw_confidence'] = np.sqrt(df_scored['confidence']) * Weights['TC_CONFIDENCE']

    # Normalize calScore from 0-1000 to 0-1 scale, then apply weight
    df_scored['cal_raw_score'] = (df_scored['calScore'] / 1000.0) * Weights['CAL_SCORE_WEIGHT']

    # ── TC Threat Score (ThreatConnect's built-in assessment) ────────────────────
    TC_THREAT_COL = 'threatAssessScore'

    if TC_THREAT_COL in df_scored.columns:
        df_scored[TC_THREAT_COL] = pd.to_numeric(
            df_scored.get(TC_THREAT_COL, 0), errors='coerce'
        ).fillna(0).clip(0, 1000)
        df_scored['tc_threat_raw_score'] = (df_scored[TC_THREAT_COL] / 1000.0) * Weights['TC_THREAT_SCORE_WEIGHT']
    else:
        df_scored['tc_threat_raw_score'] = 0.0

    # ── Incidents / Events association bonus ───────────────────────
    INCIDENTS_EVENTS_COL = 'incidents/events'

    def has_incident_event(val):
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return False
        if isinstance(val, (list, set, tuple)):
            return len(val) > 0
        s = str(val).strip()
        if s == '' or s.lower() in {'none', 'nan'}:
            return False
        return True

    if INCIDENTS_EVENTS_COL in df_scored.columns:
        df_scored['incidents_events_flag'] = df_scored[INCIDENTS_EVENTS_COL].apply(has_incident_event).astype(int)
    else:
        df_scored['incidents_events_flag'] = 0

    df_scored['incidents_events_score'] = df_scored['incidents_events_flag'] * Weights['INCIDENTS_EVENTS_WEIGHT']

    # ── Sources Count (multi-source validation) ─────────────────────
    def count_sources(sources_val):
        if sources_val is None or (isinstance(sources_val, float) and pd.isna(sources_val)):
            return 0
        if isinstance(sources_val, str):
            sources_list = [s.strip() for s in sources_val.split(',') if s.strip()]
            return len(set(sources_list))
        if isinstance(sources_val, (list, set, tuple)):
            sources_list = [str(s).strip() for s in sources_val if str(s).strip()]
            return len(set(sources_list))
        return 0

    if 'sources' in df_scored.columns:
        df_scored['sources_count'] = df_scored['sources'].apply(count_sources)
    else:
        df_scored['sources_count'] = 1

    df_scored['sources_count_safe'] = df_scored['sources_count'].clip(lower=1)
    df_scored['sources_raw_score'] = np.log1p(df_scored['sources_count_safe'] - 1) * Weights['SOURCES_WEIGHT']

    # ── Partners Count (bonus for # of partners) ─────────────────────
    MAX_PARTNERS_REALISTIC = 10

    def count_partners(partners_val):
        if partners_val is None or (isinstance(partners_val, float) and pd.isna(partners_val)):
            return 0
        if isinstance(partners_val, (list, set, tuple)):
            items = [str(x).strip() for x in partners_val if str(x).strip()]
            return len(set(items))
        s = str(partners_val).strip()
        if s == '' or s.lower() in {'none', 'nan'}:
            return 0
        lst = convert_to_list(s)
        items = [str(x).strip() for x in lst if str(x).strip()]
        return len(set(items))

    if 'partners' in df_scored.columns:
        df_scored['partners_count'] = df_scored['partners'].apply(count_partners)
    else:
        df_scored['partners_count'] = 0

    df_scored['partners_count_safe'] = df_scored['partners_count'].clip(lower=1)
    df_scored['partner_raw_score'] = np.log1p(df_scored['partners_count_safe'] - 1) * Weights['PARTNER_WEIGHT']

    # ── Threat Actor detection ───────────────────────────────────────
    def has_threat_actor(val):
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return False
        if isinstance(val, str):
            s = str(val).strip()
            if s == '' or s.lower() in {'none', 'nan'}:
                return False
            return True
        if isinstance(val, (list, set, tuple)):
            return len(val) > 0
        return False

    threat_actor_flag = pd.Series(False, index=df_scored.index)

    if 'adversary' in df_scored.columns:
        threat_actor_flag = df_scored['adversary'].apply(has_threat_actor)
    elif 'threat_actor' in df_scored.columns:
        threat_actor_flag = df_scored['threat_actor'].apply(has_threat_actor)

    threat_actor_flag = threat_actor_flag.astype(int)
    df_scored['threat_actor_score'] = threat_actor_flag * Weights['THREAT_ACTOR_WEIGHT']

    # ── TOR Activity detection ───────────────────
    def has_tor_activity(val):
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return False
        if isinstance(val, (list, set, tuple)):
            if len(val) == 0:
                return False
            t = " ".join(map(str, val)).lower()
        elif isinstance(val, str):
            if not val or val.strip() == '':
                return False
            t = " ".join(x.strip() for x in val.split(',')).lower()
        else:
            t = str(val).lower()
            if t in ['nan', 'none', '']:
                return False
        return any(keyword in t for keyword in TOR_ACTIVITY)

    tor_mask_tag = pd.Series(False, index=df_scored.index)
    tor_mask_enrich_tags = pd.Series(False, index=df_scored.index)

    if 'enrich_tags' in df_scored.columns:
        tor_mask_enrich_tags = df_scored['enrich_tags'].apply(has_tor_activity)

    if 'tag_name' in df_scored.columns:
        tag_name_as_list = df_scored['tag_name'].apply(convert_to_list)
        tor_mask_tag = tag_name_as_list.apply(has_tor_activity)

    tor_flag = (tor_mask_enrich_tags | tor_mask_tag).astype(int)
    df_scored['tor_activity_score'] = tor_flag * Weights['TOR_ACTIVITY_WEIGHT']

    # Boost TOR contribution when VT is present and high (>=10)
    vt_series_present = pd.to_numeric(df_scored['vt_numeric_for_scoring'], errors='coerce').fillna(0)
    boost_mask = (df_scored['vt_present']) & (vt_series_present >= 10) & tor_flag.astype(bool)
    df_scored.loc[boost_mask, 'tor_activity_score'] = df_scored.loc[boost_mask, 'tor_activity_score'] * 2

    # ── Stacked context bonus ─────────────────────
    # Rewards indicators with multiple reinforcing threat signals
    df_scored['stacked_context_count'] = (
        (df_scored['threat_actor_score'] > 0).astype(int) +
        (df_scored['tor_activity_score'] > 0).astype(int) +
        (df_scored['incidents_events_score'] > 0).astype(int) +
        (df_scored['sources_count'] >= 2).astype(int) +
        (df_scored['partners_count'] >= 2).astype(int)
    )

    df_scored['stacked_context_bonus'] = np.select(
        [
            df_scored['stacked_context_count'] >= 4,
            df_scored['stacked_context_count'] == 3,
            df_scored['stacked_context_count'] == 2,
        ],
        [25.0, 15.0, 7.0],
        default=0.0
    )

    # Raw additive score (before penalties/multipliers)
    df_scored['raw_score'] = (
        df_scored['malicious_raw_score'] +
        df_scored['continuity_raw_score'] +
        df_scored['tc_raw_rating'] +
        df_scored['tc_raw_confidence'] +
        df_scored['tor_activity_score'] +
        df_scored['incidents_events_score'] +
        df_scored['sources_raw_score'] +
        df_scored['partner_raw_score'] +
        df_scored['threat_actor_score'] +
        df_scored['cal_raw_score'] +
        df_scored['tc_threat_raw_score'] +
        df_scored['first_obs_raw_score'] +
        df_scored['stacked_context_bonus']
    )

    # ── Observation penalty (multiplier; linear) ────────────────────
    OBS_PENALTY_STRENGTH = float(Weights['OBSERVATION_COUNT_WEIGHT'])
    OBS_MIN_MULTIPLIER = 0.50

    obs_frac = df_scored['obs_count'] / MAX_OBS_REALISTIC
    df_scored['obs_penalty_multiplier'] = (1.0 - OBS_PENALTY_STRENGTH * obs_frac).clip(OBS_MIN_MULTIPLIER, 1.0)
    df_scored['raw_score'] *= df_scored['obs_penalty_multiplier']

    # ── Data quality multiplier (light guard) ───────────────────────
    needed = ['type', 'rating', 'confidence']
    present_frac = df_scored[needed].notna().sum(axis=1) / len(needed)
    df_scored['data_quality_multiplier'] = present_frac.clip(DATA_QUALITY_FLOOR, 1.0)
    df_scored['raw_score'] *= df_scored['data_quality_multiplier']

    # ── Botnet flag ────────────────────────────────────────────────
    col = df_scored.get('Botnet', None)
    if col is None:
        df_scored['botnet_flag'] = 0
    else:
        def is_botnet(val):
            text = " ".join(map(str, val)).lower() if isinstance(val, (list, set, tuple)) else str(val).lower()
            return int(any(action in text for action in BOTNET_ACTIONS))
        df_scored['botnet_flag'] = pd.Series(col).apply(is_botnet).astype(int)

    # ── False Positive penalty ─────────────────────────────────────
    if 'falsePositives' in df_scored.columns:
        df_scored['falsePositives'] = pd.to_numeric(df_scored['falsePositives'], errors='coerce').fillna(0)
        mask_fp = df_scored['falsePositives'] > 0
        df_scored['false_positive_raw_score'] = df_scored['raw_score'] * FALSE_POSITIVE_WEIGHT
        df_scored.loc[mask_fp, 'raw_score'] = df_scored.loc[mask_fp, 'false_positive_raw_score']
    else:
        df_scored['falsePositives'] = 0
        df_scored['false_positive_raw_score'] = df_scored['raw_score']

    # ── Scanner penalty ────────────────────────────────────────────
    def has_scanner_tag(val):
        scanners = {
            'scanner', 'masscan', 'zmap', 'shodan', 'censys',
            'active scanning: scanning ip blocks', 'web scanner', 'active scanning'
        }
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return False
        if isinstance(val, (list, set, tuple)):
            if len(val) == 0:
                return False
            t = " ".join(map(str, val)).lower()
        elif isinstance(val, str):
            if not val or val.strip() == '':
                return False
            t = " ".join(x.strip() for x in val.split(',')).lower()
        else:
            t = str(val).lower()
            if t in ['nan', 'none', '']:
                return False
        return any(s in t for s in scanners)

    scanner_mask_enrich = pd.Series(False, index=df_scored.index)
    scanner_mask_tag = pd.Series(False, index=df_scored.index)

    if 'enrich_tags' in df_scored.columns:
        scanner_mask_enrich = df_scored['enrich_tags'].apply(has_scanner_tag)

    if 'tag_name' in df_scored.columns:
        tag_name_as_list_for_scanner = df_scored['tag_name'].apply(convert_to_list)
        scanner_mask_tag = tag_name_as_list_for_scanner.apply(has_scanner_tag)

    scanner_mask = scanner_mask_enrich | scanner_mask_tag
    scanner_mask = scanner_mask & ~df_scored['is_file_type']

    df_scored['scanner_penalty_multiplier'] = np.where(scanner_mask, SCANNER_PENALTY_MULTIPLIER, 1.0)
    df_scored['raw_score'] *= df_scored['scanner_penalty_multiplier']

    # ── Botnet penalty application ──────────────────────────────────
    botnet_mask_final = (df_scored['botnet_flag'] == 1) & ~df_scored['is_file_type']
    df_scored['botnet_penalty_multiplier'] = np.where(botnet_mask_final, BOTNET_MULTIPLIER, 1.0)
    df_scored['raw_score'] *= df_scored['botnet_penalty_multiplier']

    # ── Mass Scanner penalty (Tiered) ────────────────────────────────
    # Tier 1 / Tier 2: raw_score multipliers only (no PRISM score ceiling).
    # Tier 1 (10K–100K obs, 5+ OpDivs): 60% raw score reduction.
    # Tier 2 (100K+ obs, 5+ OpDivs): 95% raw score reduction.
    MASS_SCANNER_TIER1_MULTIPLIER = 0.40   # 60% reduction
    MASS_SCANNER_TIER2_MULTIPLIER = 0.05   # 95% reduction

    tier1_col = df_scored.get('mass_scanner_tier1', None)
    tier2_col = df_scored.get('mass_scanner_tier2', None)

    mass_scanner_tier1_mask = (
        tier1_col.astype(bool) & ~df_scored['is_file_type']
        if tier1_col is not None
        else pd.Series(False, index=df_scored.index)
    )
    mass_scanner_tier2_mask = (
        tier2_col.astype(bool) & ~df_scored['is_file_type']
        if tier2_col is not None
        else pd.Series(False, index=df_scored.index)
    )

    df_scored['mass_scanner_penalty_multiplier'] = np.select(
        [mass_scanner_tier2_mask, mass_scanner_tier1_mask],
        [MASS_SCANNER_TIER2_MULTIPLIER, MASS_SCANNER_TIER1_MULTIPLIER],
        default=1.0
    )
    df_scored['raw_score'] *= df_scored['mass_scanner_penalty_multiplier']

    n1 = mass_scanner_tier1_mask.sum()
    n2 = mass_scanner_tier2_mask.sum()
    if n1 > 0:
        print(f"Mass scanner Tier 1 penalty (60% reduction) applied to {n1} indicators.")
    if n2 > 0:
        print(f"Mass scanner Tier 2 penalty (95% reduction) applied to {n2} indicators.")

    # ── Absolute Cap (only additive parts) ──────────────────────────
    MAX_SOURCES_REALISTIC = 8
    MAX_PARTNERS_REALISTIC = 10
    MAX_STACKED_CONTEXT_BONUS = 25.0

    BASE_CAP = (
        np.power(VT_EFFECTIVE_MAX, MALICIOUS_EXPONENT) * Weights['MALICIOUS_WEIGHT']+
        3 * Weights['CONTINUITY_WEIGHT'] +
        (MAX_RATING * Weights['TC_RATING']) +
        (np.sqrt(MAX_CONFIDENCE) * Weights['TC_CONFIDENCE']) +
        (Weights['TOR_ACTIVITY_WEIGHT'] * 2) +
        Weights['INCIDENTS_EVENTS_WEIGHT'] +
        (np.log1p(MAX_SOURCES_REALISTIC - 1) * Weights['SOURCES_WEIGHT']) +
        (np.log1p(MAX_PARTNERS_REALISTIC - 1) * Weights['PARTNER_WEIGHT']) +
        Weights['THREAT_ACTOR_WEIGHT'] +
        Weights['CAL_SCORE_WEIGHT'] +
        Weights['TC_THREAT_SCORE_WEIGHT'] +
        Weights['FIRST_OBS_WEIGHT'] +
        MAX_STACKED_CONTEXT_BONUS
    )

    FILE_BASELINE_RAW = 900.0

    df_scored['raw_score_cap_row'] = np.where(
        df_scored['is_file_type'],
        BASE_CAP + FILE_BASELINE_RAW,
        BASE_CAP
    )

    # ── Normalize to 0–1000 ─────────────────────────────────────────
    df_scored['PRISM_Score'] = (
        1000 * (df_scored['raw_score'] / df_scored['raw_score_cap_row']).clip(0, 1)
    )

    # Optional slight upward calibration to help uncover more high/critical indicators
    df_scored['PRISM_Score'] = np.minimum(df_scored['PRISM_Score'] * 1.40, 1000).round().fillna(0).astype(int)

    # ── VT-driven score ceilings / floors (ONLY when VT is present) ──
    vt_present_mask = df_scored['vt_present']
    vt_counts_present = df_scored['vt_numeric_for_scoring']

    low_cap_mask = vt_present_mask & (vt_counts_present <= 3)
    high_floor_mask = vt_present_mask & (vt_counts_present >= 13)

    tor_exception_mask = boost_mask

    low_max_score = scoring_bins[1] - 1
    low_cap_final_mask = low_cap_mask & ~tor_exception_mask
    df_scored.loc[low_cap_final_mask, 'PRISM_Score'] = df_scored.loc[low_cap_final_mask, 'PRISM_Score'].clip(upper=low_max_score)

    # Keep your existing behavior: floor to 499, not 500
    medium_max_score = scoring_bins[2] - 1
    df_scored.loc[high_floor_mask, 'PRISM_Score'] = df_scored.loc[high_floor_mask, 'PRISM_Score'].clip(lower=medium_max_score)

    # ── Enhanced Botnet cap with TOR exception ───────────────────────
    if 'botnet_flag' in df_scored.columns:
        botnet_cap_final_mask = (df_scored['botnet_flag'] == 1) & ~tor_exception_mask & ~df_scored['is_file_type']
        df_scored.loc[botnet_cap_final_mask, 'PRISM_Score'] = df_scored.loc[botnet_cap_final_mask, 'PRISM_Score'].clip(upper=low_max_score)

    df_scored['PRISM_Score'] = df_scored['PRISM_Score'].round().astype(int)

    # ── Assign Severity bin ─────────────────────────────────────────
    df_scored['Severity'] = pd.cut(df_scored['PRISM_Score'], bins=scoring_bins, labels=label_bins, right=False)

    # ── File Hash Floor: Never drop below CRITICAL ─────────────────
    file_hash_mask = df_scored['is_file_type']
    critical_floor = scoring_bins[3]
    df_scored.loc[file_hash_mask, 'PRISM_Score'] = df_scored.loc[file_hash_mask, 'PRISM_Score'].clip(lower=critical_floor)
    df_scored.loc[file_hash_mask, 'Severity'] = 'critical'

    # ── Explanation (drivers + multipliers) ─────────────────────────
    _NAME_MAP = {
        'malicious_raw_score': 'VT malicious (log-scaled)',
        'continuity_raw_score': 'Continuity (indicator type)',
        'tc_raw_rating': 'TC rating',
        'tc_raw_confidence': 'TC confidence',
        'tor_activity_score': 'TOR activity',
        'incidents_events_score': 'Incident/Event association',
        'sources_raw_score': 'Multi-source validation',
        'partner_raw_score': 'Partner coverage bonus',
        'threat_actor_score': 'Threat actor association',
        'cal_raw_score': 'CAL score',
        'tc_threat_raw_score': 'TC threat assessment',
        'first_obs_raw_score': 'Recent first-seen activity',
        'stacked_context_bonus': 'Reinforcing context bonus',
    }

    def build_explanation(row):
        from datetime import datetime, UTC

        parts = {
            'malicious_raw_score': float(row.get('malicious_raw_score', 0) or 0),
            'continuity_raw_score': float(row.get('continuity_raw_score', 0) or 0),
            'tc_raw_rating': float(row.get('tc_raw_rating', 0) or 0),
            'tc_raw_confidence': float(row.get('tc_raw_confidence', 0) or 0),
            'tor_activity_score': float(row.get('tor_activity_score', 0) or 0),
            'incidents_events_score': float(row.get('incidents_events_score', 0) or 0),
            'sources_raw_score': float(row.get('sources_raw_score', 0) or 0),
            'partner_raw_score': float(row.get('partner_raw_score', 0) or 0),
            'threat_actor_score': float(row.get('threat_actor_score', 0) or 0),
            'cal_raw_score': float(row.get('cal_raw_score', 0) or 0),
            'tc_threat_raw_score': float(row.get('tc_threat_raw_score', 0) or 0),
            'first_obs_raw_score': float(row.get('first_obs_raw_score', 0) or 0),
            'stacked_context_bonus': float(row.get('stacked_context_bonus', 0) or 0),
        }

        final = row.get('PRISM_Score_Final')
        score = float(final) if pd.notna(final) else float(row.get('PRISM_Score', 0) or 0)
        sev_final = row.get('Severity_Final')
        sev = str(sev_final) if pd.notna(sev_final) else str(row.get('Severity', 'nan'))
        current_date = datetime.now(UTC).strftime('%Y-%m-%d')

        vt_note = (
            f"VT score: {int(row.get('vt_numeric_for_scoring', 0))}."
            if bool(row.get('vt_present', False))
            else "VT score not available (neutral)."
        )

        contrib = sorted(parts.items(), key=lambda kv: abs(kv[1]), reverse=True)[:3]
        driver_bits = [(_NAME_MAP.get(k, k)) for k, v in contrib if v != 0]

        threat_actor_val = row.get('adversary')
        if pd.isna(threat_actor_val) or str(threat_actor_val).strip().lower() in {'none', 'nan', ''}:
            threat_actor_val = row.get('threat_actor')

        threat_actor_names = None
        if threat_actor_val is not None and str(threat_actor_val).strip().lower() not in {'none', 'nan', ''}:
            threat_actor_names = str(threat_actor_val).strip()
            driver_bits = [
                f"Associated with {threat_actor_names}" if bit == "Threat actor association" else bit
                for bit in driver_bits
            ]

        botnet_mult = float(row.get('botnet_penalty_multiplier', 1.0))
        scanner_mult = float(row.get('scanner_penalty_multiplier', 1.0))
        fp_cnt = int(row.get('falsePositives', 0) or 0)
        tor_score = float(row.get('tor_activity_score', 0) or 0)
        inc_flag = int(row.get('incidents_events_flag', 0) or 0)

        src_count = int(row.get('sources_count', 1) or 1)
        partner_count = int(row.get('partners_count', 0) or 0)
        stack_count = int(row.get('stacked_context_count', 0) or 0)

        botnet_note = "Botnet penalty applied." if botnet_mult < 1.0 else "No botnet penalty."
        fp_note = f"{fp_cnt} false positive tag(s)." if fp_cnt > 0 else "No false positive tags."
        scan_note = "Scanner penalty applied." if scanner_mult < 1.0 else "No scanner penalty."
        tor_note = "TOR activity detected." if tor_score > 0 else "No TOR activity."

        if src_count > 1:
            sources_val = str(row.get('sources', '')).strip()
            if sources_val and sources_val.lower() not in {'none', 'nan', ''}:
                src_note = f"Observed by {src_count} sources: {sources_val}."
            else:
                src_note = f"Observed by {src_count} sources."
        else:
            src_note = "Single source observation."

        partner_note = f"Observed across {partner_count} partner(s)." if partner_count > 0 else "No partner attribution."

        if inc_flag == 1:
            inc_events = str(row.get('incidents/events', '')).strip()
            if inc_events and inc_events.lower() not in {'none', 'nan', ''}:
                inc_note = f"Linked to incident/event: {inc_events}."
            else:
                inc_note = "Linked to incident/event."
        else:
            inc_note = "No incident/event link."

        stack_note = f"{stack_count} reinforcing context signal(s)." if stack_count > 0 else "No reinforcing context stack."

        drivers_text = "; ".join(driver_bits) if driver_bits else "No significant drivers"

        actor_sentence = (
            f" Associated threat actor(s): {threat_actor_names}."
            if threat_actor_names else ""
        )

        first_obs_score = float(row.get('first_obs_raw_score', 0) or 0)
        firstseen_val = row.get('firstseen_date')
        if first_obs_score > 0 and firstseen_val is not None and pd.notna(firstseen_val):
            fs_dt = pd.to_datetime(firstseen_val, errors='coerce')
            fs_str = fs_dt.strftime('%Y-%m-%d') if pd.notna(fs_dt) else str(firstseen_val)[:10]
            first_obs_sentence = f" First Observed boost applied [{fs_str}]."
        else:
            first_obs_sentence = ""

        ms_mult = float(row.get('mass_scanner_penalty_multiplier', 1.0) or 1.0)
        tobs = row.get('total_obs_7d')
        if ms_mult < 1.0:
            if pd.notna(tobs):
                mass_scanner_obs_note = (
                    f" Mass scanner penalty applied ({int(tobs):,} observations in the last 7 days)."
                )
            else:
                mass_scanner_obs_note = (
                    " Mass scanner penalty applied (7-day observation total not available)."
                )
        else:
            mass_scanner_obs_note = ""

        return (
            f"[{current_date}] Severity: {sev}. {vt_note} Contextual Drivers: {drivers_text}. "
            f"{partner_note} {src_note} {stack_note} {botnet_note} {fp_note} {scan_note} {tor_note} {inc_note}"
            f"{actor_sentence}{first_obs_sentence}{mass_scanner_obs_note} "
            f"Score: {score:.0f}/1000."
        )

    # -------------------------------------------------------------------
    # 2. AI Layer: Train a model to learn your scoring behavior
    # -------------------------------------------------------------------
    from sklearn.model_selection import train_test_split
    from sklearn.ensemble import HistGradientBoostingRegressor
    from sklearn.metrics import mean_absolute_error, r2_score

    feature_cols = [
        'vt_numeric_for_scoring',
        'obs_count',
        'rating',
        'confidence',
        'calScore',
        'threatAssessScore',
        'tor_activity_score',
        'incidents_events_flag',
        'sources_count',
        'partners_count',
        'threat_actor_score',
        'first_obs_raw_score',
        'stacked_context_bonus',
        'botnet_flag',
        'falsePositives',
        'scanner_penalty_multiplier',
        'data_quality_multiplier',
    ]

    for col_name in feature_cols:
        if col_name not in df_scored.columns:
            df_scored[col_name] = 0

    X = df_scored[feature_cols].copy().fillna(0)
    y = df_scored['PRISM_Score'].astype(float)

    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

    ai_model = HistGradientBoostingRegressor(max_depth=6, learning_rate=0.05, max_iter=300)
    ai_model.fit(X_train, y_train)

    y_pred = ai_model.predict(X_test)
    print("AI model MAE vs rule score:", mean_absolute_error(y_test, y_pred))
    print("AI model R^2 vs rule score:", r2_score(y_test, y_pred))

    # -------------------------------------------------------------------
    # 3. Apply AI model and build hybrid score
    # -------------------------------------------------------------------
    df_scored['PRISM_Score_AI'] = np.clip(ai_model.predict(X), 0, 1000).round().astype(int)

    RULE_WEIGHT = 0.7
    AI_WEIGHT = 0.3

    df_scored['PRISM_Score_Final'] = np.round(
        RULE_WEIGHT * df_scored['PRISM_Score'] + AI_WEIGHT * df_scored['PRISM_Score_AI']
    ).astype(int)

    # Bypass AI adjustment for file types and Tier 2 mass scanners —
    # both have deterministic rule-based outcomes that AI blending would only corrupt.
    file_type_mask = df_scored['is_file_type']
    df_scored.loc[file_type_mask, 'PRISM_Score_Final'] = df_scored.loc[file_type_mask, 'PRISM_Score']

    tier2_bypass_mask = mass_scanner_tier2_mask
    df_scored.loc[tier2_bypass_mask, 'PRISM_Score_Final'] = df_scored.loc[tier2_bypass_mask, 'PRISM_Score']

    df_scored['Severity_Final'] = pd.cut(
        df_scored['PRISM_Score_Final'],
        bins=scoring_bins,
        labels=label_bins,
        right=False
    )

    df_scored['AI_Adjustment'] = df_scored['PRISM_Score_Final'] - df_scored['PRISM_Score']

    df_scored['Explanation'] = df_scored.apply(build_explanation, axis=1)

    # -------------------------------------------------------------------
    # 4. Clean up, de-duplicate, and rename for export / reporting
    # -------------------------------------------------------------------
    if 'indicator' in df_scored.columns:
        df_scored.drop_duplicates(subset='indicator', inplace=True)

    column_rename_map = {
        'indicator': 'Indicator',
        'type': 'Indicator Type',
        'lastObserved': 'Last Observed',
        'vt_display': 'VT Display',
        'obs_count': 'Observation Yearly Count',
        'rating': 'ThreatConnect Rating',
        'obs_penalty_multiplier': 'Observation Penalty Multiplier',
        'botnet_flag': 'Botnet Flag',
        'falsePositives': 'False Positives',
        'partners': 'Partners',
        'partners_count': 'Partner Count',
        'sources_count': 'Source Count',
        'adversary': 'Adversary',
        'threat_actor': 'Threat Actor',
        'calScore': 'CAL Score',
        'threatAssessScore': 'ThreatConnect Score',
        'PRISM_Score': 'PRISM Score',
        'PRISM_Score_AI': 'PRISM Score (AI)',
        'PRISM_Score_Final': 'PRISM Score (Final)',
        'Severity': 'Severity',
        'Severity_Final': 'Severity (Final)',
        'Explanation': 'Explanation',
    }

    df_scored.rename(columns=column_rename_map, inplace=True)

    display_columns = [
        'Indicator', 'Indicator Type', 'Last Observed',
        'VT Display',
        'Observation Yearly Count',
        'ThreatConnect Rating',
        'Observation Penalty Multiplier',
        'Partner Count', 'Source Count',
        'Botnet Flag', 'False Positives', 'Partners', 'incidents/events', 'ThreatConnect Score',
        'CAL Score',
        'Adversary', 'Threat Actor',
        'PRISM Score', 'Severity', 'Explanation',
        'PRISM Score (AI)', 'PRISM Score (Final)', 'Severity (Final)', 'AI_Adjustment',
    ]

    display_columns = [c for c in display_columns if c in df_scored.columns]

    try:
        display(df_scored[display_columns])
    except NameError:
        print(df_scored[display_columns].head())


    df_scored[df_scored['Indicator'] == '92.118.39.56']

    import numpy as np
    import pandas as pd

    # --- Rule-layer importance derived from df_scored ---
    # These are the additive rule components already stored per row
    rule_components = [
        'malicious_raw_score',
        'continuity_raw_score',
        'tc_raw_rating',
        'tc_raw_confidence',
        'tor_activity_score',
        'incidents_events_score',
        'sources_raw_score',
        'partner_raw_score',
        'threat_actor_score',
        'cal_raw_score',
        'tc_threat_raw_score',
        'first_obs_raw_score',
    ]

    existing_rule_components = [c for c in rule_components if c in df_scored.columns]

    # Per-row absolute contributions
    rule_abs = df_scored[existing_rule_components].abs()
    row_sums = rule_abs.sum(axis=1).replace(0, np.nan)
    rule_rel = rule_abs.div(row_sums, axis=0)

    # Global average relative importance across all rows
    rule_global_importance = rule_rel.mean().sort_values(ascending=False)

    print("Average relative importance of rule components (across all indicators):")
    display(rule_global_importance)

    # Optional: show top drivers for a single example indicator by index
    example_idx = 0  # change to inspect a different row
    print(f"\nTop rule drivers for index {example_idx} (absolute contribution):")
    example_contrib = rule_abs.loc[example_idx].sort_values(ascending=False)
    display(example_contrib)


    # Save the DataFrame to Excel with only the columns displayed in cell 14
    import os
    import pandas as pd
    from datetime import datetime
    from openpyxl.styles import PatternFill

    output_dir = r"Z:\HTOC\Data_Analytics\Data\Threat Assessment Scores"
    os.makedirs(output_dir, exist_ok=True)

    # Create filename
    excel_filename = "Threat_Assessment_Scores.xlsx"
    excel_path = os.path.join(output_dir, excel_filename)

    # Select only the columns to save (using renamed column names)
    columns_to_save = ['Indicator', 'Last Observed', 'Indicator Type', 'VirusTotal Malicious Score', 
                       'Observation Yearly Count', 'ThreatConnect Rating', 'Observation Penalty Multiplier',
                       'Botnet Flag', 'False Positives', 'Partners', 'incidents/events', 'Threat Actor', 'CAL Score', 'ThreatConnect Score',
                       'PRISM Score', 'Severity', 'Explanation']

    # Make a copy with only selected columns (only include columns that exist)
    columns_to_save = [col for col in columns_to_save if col in df_scored.columns]
    df_export = df_scored[columns_to_save].copy()

    # Use 'PRISM Score (Final)' values for the 'PRISM Score' column in Excel
    if 'PRISM Score (Final)' in df_scored.columns:
        df_export['PRISM Score'] = df_scored['PRISM Score (Final)']
        # Also update Severity to use Severity (Final) if available
        if 'Severity (Final)' in df_scored.columns:
            df_export['Severity'] = df_scored['Severity (Final)']

    # Convert timezone-aware datetime columns to timezone-naive for Excel compatibility
    for col in df_export.columns:
        if pd.api.types.is_datetime64_any_dtype(df_export[col]):
            # Check if the column has timezone info
            if df_export[col].dt.tz is not None:
                # Convert to UTC first, then remove timezone info to make it timezone-naive
                df_export[col] = df_export[col].dt.tz_convert('UTC').dt.tz_localize(None)

    # Check if file exists and read existing data
    if os.path.exists(excel_path):
        df_existing = pd.read_excel(excel_path, engine='openpyxl')

        # Normalize existing columns to the latest schema (older files may still use raw names)
        rename_map_existing = {
            old: new for old, new in column_rename_map.items()
            if old in df_existing.columns and new not in df_existing.columns
        }
        if rename_map_existing:
            df_existing.rename(columns=rename_map_existing, inplace=True)

        # Add missing columns to existing dataframe with default values
        for col in columns_to_save:
            if col not in df_existing.columns:
                # Set appropriate default values based on column type
                if col == 'Last Observed':
                    df_existing[col] = pd.NaT
                elif col in ['VirusTotal Malicious Score', 'Observation Yearly Count', 'ThreatConnect Rating', 
                            'Observation Penalty Multiplier', 'Botnet Flag', 'False Positives', 'PRISM Score', 'ThreatConnect Score']:
                    df_existing[col] = 0
                elif col == 'Severity':
                    df_existing[col] = ''
                else:
                    df_existing[col] = ''

        # Ensure both dataframes have the same columns in the same order
        df_existing = df_existing[columns_to_save].copy()

        # Count how many indicators will be updated vs added
        existing_indicators = set(df_existing['Indicator'].values)
        new_indicators = set(df_export['Indicator'].values)

        indicators_to_add = len(new_indicators - existing_indicators)
        indicators_to_check = existing_indicators & new_indicators

        # Find indicators that have actually changed
        indicators_to_update = []
        indicators_unchanged = []

        # Set indicator as index for comparison
        df_existing_idx = df_existing.set_index('Indicator').sort_index()
        df_export_idx = df_export.set_index('Indicator').sort_index()

        for indicator in indicators_to_check:
            # Compare the rows (excluding the index/indicator column)
            existing_row = df_existing_idx.loc[indicator]
            new_row = df_export_idx.loc[indicator]

            # Check if any values have changed
            if not existing_row.equals(new_row):
                indicators_to_update.append(indicator)
            else:
                indicators_unchanged.append(indicator)

        # Build the combined dataframe: keep existing unchanged records, update changed ones, add new ones
        # Start with existing records that are unchanged
        df_unchanged = df_existing[df_existing['Indicator'].isin(indicators_unchanged)].copy()

        # Add updated records (from new data)
        df_updated = df_export[df_export['Indicator'].isin(indicators_to_update)].copy()

        # Add new records (not in existing)
        df_new = df_export[df_export['Indicator'].isin(new_indicators - existing_indicators)].copy()

        # Add existing records that are not in new data (preserve historical data)
        df_preserved = df_existing[~df_existing['Indicator'].isin(new_indicators)].copy()

        # Combine all parts
        df_combined = pd.concat([df_unchanged, df_updated, df_new, df_preserved], ignore_index=True)

        # Final check: remove any duplicates (shouldn't happen, but safety check)
        df_combined = df_combined.drop_duplicates(subset='Indicator', keep='last')

        print(f"Existing indicators in new data: {len(indicators_to_check)}")
        print(f"Added (new) indicators: {indicators_to_add}")
        print(f"Updated existing indicators: {len(indicators_to_update)}")
        print(f"Total indicators in sheet: {len(df_combined)}")
    else:
        df_combined = df_export.copy()
        # Remove any duplicates in the new data itself
        df_combined = df_combined.drop_duplicates(subset='Indicator', keep='last')
        print(f"Created new file with {len(df_combined)} indicators")

    # Read existing Complete History if it exists (to preserve it when rewriting file)
    df_complete_history = None
    if os.path.exists(excel_path):
        try:
            df_complete_history = pd.read_excel(excel_path, sheet_name='Complete History', engine='openpyxl')
            print(f"Preserving existing Complete History with {len(df_complete_history)} records")
        except:
            pass  # Sheet doesn't exist yet, that's okay

    # Save to Excel with severity-based highlighting
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df_combined.to_excel(writer, index=False, sheet_name='PRISM Scores')

        # Create and write comparison sheet
        if 'ThreatConnect Score' in df_combined.columns and 'PRISM Score' in df_combined.columns:
            comp_cols = ['Indicator', 'ThreatConnect Score', 'PRISM Score']
            df_comp = df_combined[comp_cols].copy()
            # Ensure numeric
            df_comp['ThreatConnect Score'] = pd.to_numeric(df_comp['ThreatConnect Score'], errors='coerce').fillna(0)
            df_comp['PRISM Score'] = pd.to_numeric(df_comp['PRISM Score'], errors='coerce').fillna(0)
            df_comp['Difference'] = df_comp['PRISM Score'] - df_comp['ThreatConnect Score']
            df_comp.to_excel(writer, index=False, sheet_name='Score Comparison')

        # Write Complete History if it exists (preserve it)
        if df_complete_history is not None:
            df_complete_history.to_excel(writer, index=False, sheet_name='Complete History')

        workbook = writer.book
        worksheet = writer.sheets['PRISM Scores']



        # Define fills for each severity
        fills = {
            'low': PatternFill(start_color='83de85', end_color='83de85', fill_type='solid'),     # light green
            'medium': PatternFill(start_color='eef084', end_color='eef084', fill_type='solid'),  # light yellow
            'high': PatternFill(start_color='f29953', end_color='f29953', fill_type='solid'),    # light orange
            'critical': PatternFill(start_color='e83f3f', end_color='e83f3f', fill_type='solid') # light red
        }

        for row_idx, severity in enumerate(df_combined['Severity'], start=2):  # start=2 to skip header
            fill = fills.get(str(severity).lower())
            if fill:
                for col_idx in range(1, len(df_combined.columns) + 1):
                    worksheet.cell(row=row_idx, column=col_idx).fill = fill

    print(f"Saved {len(df_combined)} total indicators with PRISM scoring results to {excel_path}")


    import os
    from datetime import datetime, UTC
    import pandas as pd
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl import load_workbook

    # ============================================================================
    # INDICATOR SCORING HISTORY - Similar to Credit Report History
    # Maintains a time-series record of all indicator scores and explanations in Excel
    # ============================================================================

    # Define history file path - using the same file as main Prioritized Risk and Indicator Severity Model scores
    history_dir = r"Z:\HTOC\Data_Analytics\Data\Threat Assessment Scores"
    os.makedirs(history_dir, exist_ok=True)

    # Use the main Threat_Assessment_Scores.xlsx file instead of separate file
    history_file = os.path.join(history_dir, "Threat_Assessment_Scores.xlsx")

    # Record the current scoring run (using timezone-aware UTC)
    run_timestamp = datetime.now(UTC).strftime('%Y-%m-%d %H:%M:%S')

    # Define columns to track - only essential fields
    history_columns = [
        'Scoring Date',
        'Indicator',
        'Indicator Type',
        'PRISM Score',
        'Severity',
        'Explanation'
    ]

    # Map dataframe columns to history columns
    column_mapping = {
        'Indicator': 'Indicator',
        'Indicator Type': 'Indicator Type',
        'PRISM Score (Final)': 'PRISM Score',
        'Severity (Final)': 'Severity',
        'Explanation': 'Explanation'
    }

    # Build history data for this run
    history_data = []
    for idx, row in df_scored.iterrows():
        history_row = {}
        for df_col, hist_col in column_mapping.items():
            if df_col in df_scored.columns:
                value = row[df_col]
                # Handle datetime/timestamp conversions
                if pd.api.types.is_datetime64_any_dtype(type(value)):
                    history_row[hist_col] = str(value) if pd.notna(value) else ''
                elif pd.isna(value):
                    history_row[hist_col] = ''
                else:
                    history_row[hist_col] = value

        history_row['Scoring Date'] = run_timestamp
        history_data.append(history_row)

    df_history_current = pd.DataFrame(history_data)

    # Reorder columns to match history_columns order
    df_history_current = df_history_current[[col for col in history_columns if col in df_history_current.columns]]

    # Load existing history from the Excel file if it exists
    if os.path.exists(history_file):
        try:
            # Try to read the Complete History sheet
            df_history_all = pd.read_excel(history_file, sheet_name='Complete History', engine='openpyxl')
            # Convert any timezone-aware columns to timezone-naive BEFORE concatenation
            for col in df_history_all.columns:
                if pd.api.types.is_datetime64_any_dtype(df_history_all[col]):
                    if hasattr(df_history_all[col].dtype, 'tz') and df_history_all[col].dtype.tz is not None:
                        df_history_all[col] = df_history_all[col].dt.tz_convert('UTC').dt.tz_localize(None)

            # Append new records to existing history - preserve all historical records
            # Exclude completely empty/all-NA columns before concatenation to avoid deprecation warning
            df_history_all = df_history_all.dropna(axis=1, how='all')
            df_history_current = df_history_current.dropna(axis=1, how='all')
            df_history_all = pd.concat([df_history_all, df_history_current], ignore_index=True)
        except:
            # If Complete History sheet doesn't exist, start fresh
            df_history_all = df_history_current.copy()
    else:
        df_history_all = df_history_current.copy()

    # Remove duplicates only if the same indicator was scored on the same date (keep most recent)
    # Extract date only (without time) for duplicate detection to prevent time-based duplicates
    df_history_all['Scoring Date Only'] = pd.to_datetime(df_history_all['Scoring Date']).dt.date
    df_history_all = df_history_all.drop_duplicates(subset=['Indicator', 'Scoring Date Only'], keep='last')
    # Remove the temporary column
    df_history_all = df_history_all.drop(columns=['Scoring Date Only'])

    # Sort by Indicator and Scoring Date for easier review
    df_history_all = df_history_all.sort_values(['Indicator', 'Scoring Date'], ascending=[True, False])

    # Convert timezone-aware datetime columns to timezone-naive for Excel compatibility
    df_history_all = df_history_all.copy()  # Make a copy to avoid SettingWithCopyWarning
    for col in df_history_all.columns:
        if pd.api.types.is_datetime64_any_dtype(df_history_all[col]):
            # Check if column has timezone info
            if hasattr(df_history_all[col].dtype, 'tz') and df_history_all[col].dtype.tz is not None:
                df_history_all[col] = df_history_all[col].dt.tz_convert('UTC').dt.tz_localize(None)
            elif df_history_all[col].dt.tz is not None:
                df_history_all[col] = df_history_all[col].dt.tz_convert('UTC').dt.tz_localize(None)

    # Reorder columns in complete history to match desired order
    df_history_all = df_history_all[[col for col in history_columns if col in df_history_all.columns]]

    # Read existing sheets to preserve them
    existing_sheets = {}
    if os.path.exists(history_file):
        try:
            excel_file = pd.ExcelFile(history_file, engine='openpyxl')
            for sheet_name in excel_file.sheet_names:
                if sheet_name not in ['Complete History', 'Latest Scores']:
                    existing_sheets[sheet_name] = pd.read_excel(history_file, sheet_name=sheet_name, engine='openpyxl')
            excel_file.close()
        except Exception as e:
            print(f"Warning: Could not read existing sheets: {e}")

    # Now write everything back - this mimics the pattern from the working code
    with pd.ExcelWriter(history_file, engine='openpyxl') as writer:
        # Write existing sheets first (PRISM Scores and Score Comparison)
        for sheet_name, sheet_df in existing_sheets.items():
            sheet_df.to_excel(writer, index=False, sheet_name=sheet_name)

            # If this is the PRISM Scores sheet, reapply the severity formatting
            if sheet_name == 'PRISM Scores' and 'Severity' in sheet_df.columns:
                worksheet = writer.sheets['PRISM Scores']

                # Define fills for each severity (same as original code)
                fills = {
                    'low': PatternFill(start_color='83de85', end_color='83de85', fill_type='solid'),     # light green
                    'medium': PatternFill(start_color='eef084', end_color='eef084', fill_type='solid'),  # light yellow
                    'high': PatternFill(start_color='f29953', end_color='f29953', fill_type='solid'),    # light orange
                    'critical': PatternFill(start_color='e83f3f', end_color='e83f3f', fill_type='solid') # light red
                }

                for row_idx, severity in enumerate(sheet_df['Severity'], start=2):  # start=2 to skip header
                    fill = fills.get(str(severity).lower())
                    if fill:
                        for col_idx in range(1, len(sheet_df.columns) + 1):
                            worksheet.cell(row=row_idx, column=col_idx).fill = fill

        # Write Complete History
        df_history_all.to_excel(writer, sheet_name='Complete History', index=False)

        # Format "Complete History" sheet
        ws_history = writer.sheets['Complete History']
        ws_history.column_dimensions['A'].width = 20  # Scoring Date
        ws_history.column_dimensions['B'].width = 20  # Indicator
        ws_history.column_dimensions['C'].width = 15  # Indicator Type
        ws_history.column_dimensions['D'].width = 18  # PRISM Score
        ws_history.column_dimensions['E'].width = 12  # Severity
        ws_history.column_dimensions['F'].width = 60  # Explanation

        # Add header formatting to Complete History
        for cell in ws_history[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    print(f"✓ Scoring history updated in: {history_file}")
    print(f"✓ Excel file now contains:")
    print(f"  - PRISM Scores (current scores with all details and severity formatting)")
    print(f"  - Score Comparison (PRISM vs ThreatConnect)")
    print(f"  - Complete History (time-series of all historical scores)")
    print(f"✓ Total indicators in history: {len(df_history_all['Indicator'].unique())}")
    print(f"✓ Total scoring records: {len(df_history_all)}")

    # ============================================================================
    # HISTORY LOOKUP FUNCTIONS - Query past scores like a credit report
    # ============================================================================

    def get_indicator_score_history(indicator):
        """
        Retrieve the complete scoring history for a single indicator.
        Similar to pulling a credit report for one person.
        """
        history = df_history_all[df_history_all['Indicator'] == indicator].copy()
        if history.empty:
            return None

        return history.sort_values('Scoring Date', ascending=False)

    def get_score_changes_since(days_ago=7):
        """
        Find all indicators whose scores have changed in the last N days.
        Useful for identifying trending threats.
        """
        cutoff_date = pd.Timestamp.utcnow() - pd.Timedelta(days=days_ago)

        changed_indicators = []
        for indicator in df_history_all['Indicator'].unique():
            ind_history = df_history_all[df_history_all['Indicator'] == indicator].sort_values('Scoring Date', ascending=False)

            if len(ind_history) > 1:
                recent = ind_history[pd.to_datetime(ind_history['Scoring Date']) >= cutoff_date]

                if len(recent) > 1:
                    old_score = recent.iloc[-1]['PRISM Score']
                    new_score = recent.iloc[0]['PRISM Score']

                    if pd.notna(old_score) and pd.notna(new_score) and old_score != new_score:
                        changed_indicators.append({
                            'Indicator': indicator,
                            'Type': recent.iloc[0]['Indicator Type'],
                            'Previous Score': old_score,
                            'Current Score': new_score,
                            'Change': new_score - old_score,
                            'Last Updated': recent.iloc[0]['Scoring Date']
                        })

        if changed_indicators:
            return pd.DataFrame(changed_indicators).sort_values('Change', ascending=False, key=abs)
        else:
            return pd.DataFrame()

    # Display summary statistics
    print("\n" + "="*70)
    print("SCORING HISTORY SUMMARY")
    print("="*70)

    # Count indicators with multiple records
    multi_record = df_history_all.groupby('Indicator').size()
    indicators_with_history = sum(multi_record > 1)
    print(f"Indicators with multiple scoring records: {indicators_with_history}")

    # Show example: lookup a specific indicator if any exist
    if not df_scored.empty:
        example_indicator = df_scored['Indicator'].iloc[0]
        example_history = get_indicator_score_history(example_indicator)
        if example_history is not None:
            print(f"\nExample - Scoring history for {example_indicator}:")
            # Only show columns that exist
            example_cols = ['Scoring Date']
            if 'PRISM Score' in example_history.columns:
                example_cols.append('PRISM Score')
            if 'Severity' in example_history.columns:
                example_cols.append('Severity')
            if 'Explanation' in example_history.columns:
                example_cols.append('Explanation')
            print(example_history[example_cols].head())

    print(f"\n✓ History functions available:")
    print(f"  - get_indicator_score_history(indicator)  # Get full history for one indicator")
    print(f"  - get_score_changes_since(days_ago=7)      # Find indicators with score changes")


if __name__ == "__main__":
    main()
